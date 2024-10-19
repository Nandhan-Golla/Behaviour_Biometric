import openpyxl
import time
import openpyxl.workbook
import pyautogui
"""
workbook = openpyxl.Workbook()
sheet = workbook.active
workbook.save('Data.xlsx')"""

try:
   workbook = openpyxl.load_workbook('Data.xlsx')
   sheet = workbook.active
   
except Exception:
   workbook = openpyxl.Workbook()
   sheet = workbook.active
   workbook.save('Data.xlsx')

if sheet.max_row == 1:
   m_rows = 1

else:
   m_rows = sheet.max_row+1

def calculate_typing_speed(text, time_taken):

    num_words = len(text.split())
    wpm = (num_words / time_taken) * 60
    return wpm

srt_ko = []
def execute():
    print("Enter the text you want to type:")
    text_to_type = input()

    start_time = time.time()
    print("Start typing...")
    input(":") 
    end_time = time.time()

    time_taken = end_time - start_time



    typing_speed = calculate_typing_speed(text_to_type, time_taken)
    print(f"Your typing speed is: {typing_speed:.2f} WPM")
    srt_ko.append(typing_speed)
"""
    workbook = openpyxl.Workbook()

    sheet = workbook.active
#sheet.column_dimensions['A'].letter = 'Mouse_x'
    workbook.save('BB.xlsx')"""

def fill_column(filename, column_letter, data):


    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for i, value in enumerate(data):
        cell = sheet[column_letter + str((m_rows) + i)]
        cell.value = value

    workbook.save(filename)

# Example usage:
"""filename = 'example.xlsx'
column_letter = 'A'
data = ['Value 1', 'Value 2', 'Value 3']"""

#fill_column(filename, column_letter, data)
#filename = 'BB.xls


def process(user):

  mouse_x =[]
  mouse_y = []
  filename = 'Data.xlsx'
  #global user
#fill_column(filename, 'D', str(input("Entr Profile ID: ")))


  res_factor = 1
  while True:
     execute()

     x, y = pyautogui.position()
     print(f'Mouse Position: {x} , {y}')
    #user1_mouse_movements.append((x, y))
     mouse_x.append(x)
     mouse_y.append(y)
     #filename = 'BB.xlsx'
     column_letter_x = 'A'
     column_letter_y = 'B'
     column_letter_key = 'C'
    
     fill_column(filename, column_letter_x, mouse_x)
     fill_column(filename,column_letter_y, mouse_y)
     fill_column(filename, column_letter_key, srt_ko)

     time.sleep(0.5)
     res_factor += 1
     if res_factor == 6:
        break


  workbook = openpyxl.load_workbook('Data.xlsx')

  sheet = workbook['Sheet']

  for i in range(res_factor-1):
     sheet[f'D{m_rows+i}'].value = user
     pass
  
  if sheet['A1'].value == 'Mouse_x':
     workbook.save('Data.xlsx')

  else:
       sheet.insert_rows(idx=1)
       sheet['A1'].value = 'Mouse_x'
       sheet['B1'].value = 'Mouse_y'
       sheet['C1'].value = 'Strokes'
       sheet['D1'].value = 'User'
       workbook.save('Data.xlsx')
"""
def target_func():
   s_array =[]
   s_array.append((mouse_x[0], mouse_y[0], srt_ko[0]))

     
"""





  




